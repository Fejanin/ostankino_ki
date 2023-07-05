import openpyxl


class XLSX_Data:
    '''Хранит данные по одной строке:
- код единицы продаж;
- код продукта;
- номер варианта;
- штрих-код;
- наименование;
- количество в заказе;
- и порядковый номер строки, из которой были записаны данные.'''
    START_SLICE = 0
    END_SLICE = 4
    NAME_PRODUCT = 12
    NUM_ORDER = 20
    CONTROL_KG = 19
    NAME_COL = 'U'
    def __new__(cls, data, number, *args, **kwargs):
        res = list(map(lambda x: x.value, data))
        if cls.control_line(res):
            return super().__new__(cls)


    def __init__(self, data, number):
        self.sales_unit_code, self.product_code, self.option_number, self.barcode, self.product, self.num_order = self.read_line(data)
        self.num_row = number


    def read_line(self, data):
        res = list(map(lambda x: x.value, data))
        if self.control_line(res):
            return res[self.START_SLICE: self.END_SLICE] + [res[self.NAME_PRODUCT], res[self.NUM_ORDER]]


    def __str__(self):
        return f'{self.num_row}){self.product}: {self.num_order}.'


    def __eq__(self, obj):
        if type(self) == type(obj):
            return self.sales_unit_code == obj.sales_unit_code and self.product_code == obj.product_code and self.option_number == obj.option_number and self.barcode == obj.barcode
        else:
            # TODO
            # сравнение с другими объектами
            return False


    def __hash__(self):
        '''пока не используем'''
        return hash((self.sales_unit_code, self.product_code, self.option_number, self.barcode))


    def __setattr__(self, key, value):
        '''Удаляем пробелы из начала и окончания строки.'''
        if type(value) is str:
            value = value.strip()
        object.__setattr__(self, key, value)


    @classmethod
    def control_line(cls, data):
        return all(data[cls.START_SLICE: cls.END_SLICE]) and data[cls.CONTROL_KG] == 'кг'


class POKOM_Reader:
    '''Скачивает данные из file.
Формирует список объектов XLSX_Data'''
    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb.active
        self.all_rows = []
        self.read()


    def read(self):
        a = self.ws.iter_rows(1, self.ws.max_row + 1)
        for num, i in enumerate(a, 1):
            obj = XLSX_Data(i, num)
            if obj:
                self.all_rows.append(obj)


    def __call__(self):
        return round(sum([i.num_order for i in self.all_rows]), 3)


class POKOM_Rewriter:
    def __init__(self, read_name_file, write_name_file, flag_pocom = True):
        if flag_pocom: # файл ПОКОМ
            self.read_file = POKOM_Reader(read_name_file)
        else:
            # TODO
            # выбрать клиента, для использования соответствующего класса
            pass
        self.write_file = POKOM_Reader(write_name_file)
        for i in self.read_file.all_rows:
            flag = False
            for j in self.write_file.all_rows:
                if i == j and i.num_order:
                    flag = True
                    #XLSX_Writer(self.write_file.ws, j, i.num_order)
                    self.write(j, i.num_order)
                    # TODO
                    # добавить в треккер изменения (Tracker)
                    continue
            if not flag and i.num_order: # ТРЕБУЕТСЯ ПРОВЕРИТЬ ПРАВИЛЬНОСТЬ ПРОВЕРКИ!!!
                # TODO
                # добавить в треккер ошибку (Tracker)
                print('NOT FOUND')
    # СОХРАНИТЬ изменени в файлах
        self.read_file.wb.save(read_name_file)
        self.write_file.wb.save(write_name_file)


    def write(self, obj, num):
        self.write_file.ws[obj.NAME_COL + str(obj.num_row)] = num


class XLSX_Writer:
    def __init__(self, sheet, obj_to, num_order):
        sheet[obj_to.NAME_COL + str(obj_to.num_row)] = num_order


class Tracker:
    '''Создает текстовый файл с результатами переноса данных (в т.ч. и ошибками)'''
    pass
