import openpyxl


class ReaderData:
    '''Считываем данные из заказа клиента'''
    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb.active
        self.error = []
        self.all_keys = {} # {'3215 ВЕТЧ.МЯСНАЯ Папа может п/о 0.4кг 8шт.    ОСТАНКИНО': {'order': 160, 'num_row': 1}, '3678 СОЧНЫЕ сос п/о мгс 2*2     ОСТАНКИНО': {'order': 2200, 'num_row': 3}, ...}
        self.read()


    def read(self):
        a = self.ws.iter_rows(1, self.ws.max_row)
        for num, i in enumerate(a, 1):
            if  i[1].value:
                if i[0].value in  self.all_keys:
                    self.error.append(f'{i[0].value} из строки №{num} уже ранее БЫЛО в строке №{self.all_keys[i[0].value]["num_row"]} бланка заказа.')
                    continue
                self.all_keys[i[0].value] = {'order': i[1].value, 'num_row': num}


class Translater:
    '''Создаем словарь для сопоставления данных'''
    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb.active
        self.error = []
        self.all_keys = {} # {'3215 ВЕТЧ.МЯСНАЯ Папа может п/о 0.4кг 8шт.    ОСТАНКИНО': (1001094053215, 'ВЕТЧ.МЯСНАЯ Папа может п/о 0.4кг 8шт.'), '3248 ДОКТОРСКАЯ ТРАДИЦ. вар п/о ОСТАНКИНО': (1001010113248, 'ДОКТОРСКАЯ ТРАДИЦ. вар п/о'), ...}
        self.read()


    def read(self):
        a = self.ws.iter_rows(1, self.ws.max_row)
        for i in a:
            if i[0].value and i[0].value != '1С':
                self.all_keys[i[0].value] = (i[1].value, i[2].value)


class Writer:
    '''Создаем словарь на основании бланка заказа Останкино'''
    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb.active
        self.error = []
        self.all_keys = {} # {(1001010105246, 'ДОКТОРСКАЯ ПРЕМИУМ вар б/о мгс_30с'): 11, (1001303636415, 'БАЛЫКОВАЯ Коровино п/к в/у 0.84кг 6шт.'): 12, ...}
        self.read()


    def read(self):
        a = self.ws.iter_rows(1, self.ws.max_row)
        for n, i in enumerate(a, 1):
            if i[2].value in ('КГ', 'ШТ', 'ШТ.', 'кг', 'шт', 'шт.', 'Кг', 'Шт', 'Шт.'):
                self.all_keys[(i[3].value, i[1].value)] = n


class OldSKU:
    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb.active
        self.error = []
        self.all_keys = [] # ['4571 МОЛОЧНАЯ ТРАДИЦ. вар ОСТАНКИН', '5889 ОСОБАЯ Коровино вар п/о 0.4кг 8шт.  ОСТАНКИНО', '6065 ПОСОЛЬСКАЯ с/к с/н в/у 1/100 8шт.   ОСТАНКИНО', ...]
        self.read()


    def read(self):
        a = self.ws.iter_rows(1, self.ws.max_row)
        for i in a:
            self.all_keys.append(i[0].value)





#file_data = ReaderData('test.xlsx')
#translate = Translater('1С.xlsx')
#write_to_file = Writer('бланк заказа.xlsx')
#old_sku = OldSKU('old_SKU.xlsx')
